# email_connector.py

import os
import re
import smtplib
import imaplib
import email
import tempfile
import unicodedata
import socket
from datetime import datetime
from logger import logger
from email.mime.base import MIMEBase
from email import encoders


class EmailConnector:
    """Conector genérico para servicios de correo mediante SMTP e IMAP."""

    def __init__(self, smtp_server, smtp_port, imap_server, imap_port,
                 email_address, password, use_tls=True):
        self.smtp_server = smtp_server
        self.smtp_port = int(smtp_port)
        self.imap_server = imap_server
        self.imap_port = int(imap_port)
        self.email_address = email_address
        self.password = password
        self.use_tls = use_tls
        # Configurar timeout por defecto para sockets
        socket.setdefaulttimeout(30)

    def is_available(self):
        """Verifica si las dependencias básicas están disponibles."""
        return True  # smtplib e imaplib son parte de la biblioteca estándar

    def test_connection(self):
        """Prueba la conexión SMTP con las credenciales proporcionadas."""
        try:
            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=10) as server:
                if self.use_tls:
                    server.starttls()
                server.login(self.email_address, self.password)
            return True, "Conexión SMTP exitosa"
        except Exception as e:
            logger.error(f"Error en conexión SMTP: {e}")
            return False, str(e)

    def load_folders(self, callback=None):
        """Obtiene la lista de carpetas disponibles mediante IMAP."""
        folders = []
        imap = None
        try:
            # Agregar timeout de 30 segundos para evitar bloqueos indefinidos
            imap = imaplib.IMAP4_SSL(self.imap_server, self.imap_port, timeout=30)
            imap.login(self.email_address, self.password)
            typ, data = imap.list()
            if typ == 'OK':
                for line in data:
                    decoded = line.decode()
                    parts = decoded.split(' "/" ')
                    if len(parts) == 2:
                        folder = parts[1].strip('"')
                        folders.append(folder)
                        if callback:
                            callback(f"Bandeja encontrada: {folder}")
        except Exception as e:
            if callback:
                callback(f"Error al cargar carpetas: {e}", "ERROR")
            logger.error(f"Error al cargar carpetas IMAP: {e}")
        finally:
            if imap:
                try:
                    imap.logout()
                except:
                    pass
        return folders

    @staticmethod
    def _normalize_text(value):
        normalized = unicodedata.normalize('NFKD', value or '')
        return ''.join(ch for ch in normalized if not unicodedata.combining(ch)).lower()

    @classmethod
    def _prepare_title_filters(cls, title_filter):
        if not title_filter:
            return []

        filters = []
        for raw_group in re.split(r'[;,|]+', title_filter):
            tokens = [cls._normalize_text(token) for token in raw_group.split() if token]
            if tokens:
                filters.append(tokens)
        return filters

    @classmethod
    def _subject_matches(cls, subject, prepared_filters):
        if not prepared_filters:
            return True

        normalized_subject = cls._normalize_text(subject)
        for tokens in prepared_filters:
            if all(token in normalized_subject for token in tokens):
                return True
        return False

    def search_emails_and_download_excel(self, folder_path, title_filter,
                                         today_only=True, status_callback=None,
                                         result_callback=None, max_emails_to_check=50,
                                         max_matches=10):
        """
        Busca correos por título y descarga adjuntos de Excel.
        OPTIMIZADO: Procesa correos uno a uno con límites para evitar consumo excesivo.

        Args:
            folder_path: Carpeta de correo a buscar
            title_filter: Filtro de título para buscar correos
            today_only: Si True, solo busca correos de hoy (default: True)
            status_callback: Callback para reportar estado
            result_callback: Callback para reportar resultados
            max_emails_to_check: Máximo de correos a revisar (default: 50)
            max_matches: Máximo de coincidencias a procesar (default: 10)
        """
        results = {
            "success": False,
            "total_items": 0,
            "matching_items": 0,
            "excel_files": [],
            "errors": [],
            "temp_dir": None,
            "message": ""
        }

        folder_path = (folder_path or "INBOX").strip() or "INBOX"
        prepared_filters = self._prepare_title_filters(title_filter)

        temp_dir = tempfile.mkdtemp(prefix="enlace_db_excel_")
        results["temp_dir"] = temp_dir

        imap = None
        try:
            # Agregar timeout de 30 segundos
            imap = imaplib.IMAP4_SSL(self.imap_server, self.imap_port, timeout=30)
            imap.login(self.email_address, self.password)
            imap.select(folder_path)

            date_str = datetime.now().strftime('%d-%b-%Y')
            search_criteria = ['ON', date_str] if today_only else ['ALL']

            typ, data = imap.search(None, *search_criteria)
            if typ != 'OK':
                results["message"] = "Error en búsqueda IMAP"
                return results

            email_ids = data[0].split()
            total_emails = len(email_ids)

            if not email_ids:
                results["success"] = True
                results["message"] = "No se encontraron correos"
                return results

            # OPTIMIZACIÓN: Limitar cantidad de correos a revisar
            emails_to_process = email_ids[:max_emails_to_check]

            if status_callback:
                if total_emails > max_emails_to_check:
                    status_callback(f"Encontrados {total_emails} correos. Procesando los primeros {max_emails_to_check}...", "INFO")
                else:
                    status_callback(f"Encontrados {total_emails} correos. Procesando...", "INFO")

            # Procesar cada correo UNO A UNO
            emails_checked = 0
            for num in emails_to_process:
                emails_checked += 1

                # OPTIMIZACIÓN: Detener si ya encontramos suficientes coincidencias
                if results["matching_items"] >= max_matches:
                    if status_callback:
                        status_callback(f"Límite de {max_matches} coincidencias alcanzado. Deteniendo búsqueda.", "INFO")
                    break

                # Leer solo el encabezado para obtener el asunto sin marcar como leído
                typ, header_data = imap.fetch(num, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
                if typ != 'OK':
                    continue

                header_msg = email.message_from_bytes(header_data[0][1])
                subject = header_msg.get('Subject', '')

                # Verificar si el asunto coincide con el filtro
                if not self._subject_matches(subject, prepared_filters):
                    # No coincide, continuar con el siguiente
                    if status_callback and emails_checked % 10 == 0:
                        status_callback(f"Revisados {emails_checked} correos...", "INFO")
                    continue

                # COINCIDENCIA ENCONTRADA - Descargar el mensaje completo
                typ, msg_data = imap.fetch(num, '(BODY.PEEK[])')
                if typ != 'OK':
                    continue

                msg = email.message_from_bytes(msg_data[0][1])
                results["matching_items"] += 1

                if status_callback:
                    status_callback(f"✓ Coincidencia #{results['matching_items']}: '{subject}'", "SUCCESS")

                # Buscar adjuntos Excel
                for part in msg.walk():
                    if part.get_content_disposition() == 'attachment':
                        filename = part.get_filename()
                        if filename and filename.lower().endswith(('.xls', '.xlsx')):
                            filepath = os.path.join(temp_dir, filename)
                            with open(filepath, 'wb') as f:
                                f.write(part.get_payload(decode=True))
                            results["excel_files"].append(filepath)
                            if status_callback:
                                status_callback(f"Descargado: {filename}", "SUCCESS")

                # Marcar como leído después de procesar
                imap.store(num, '+FLAGS', '\\Seen')

            results["total_items"] = emails_checked
        except socket.timeout:
            results["errors"].append("Timeout de conexión IMAP")
            if status_callback:
                status_callback("Error: Timeout de conexión IMAP", "ERROR")
            logger.error("Timeout de conexión IMAP")
        except Exception as e:
            results["errors"].append(str(e))
            if status_callback:
                status_callback(f"Error: {e}", "ERROR")
            logger.error(f"Error en búsqueda IMAP: {e}")
        finally:
            if imap:
                try:
                    imap.logout()
                except:
                    pass

        if results["errors"]:
            return results

        results["success"] = True
        results["message"] = f"{results['matching_items']} correos coincidentes"
        return results

    def send_simple_email(self, to_email, subject, body):
        """Envía un correo simple a un destinatario."""
        try:
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            msg = MIMEMultipart()
            msg['From'] = self.email_address
            msg['To'] = to_email
            msg['Subject'] = subject

            msg.attach(MIMEText(body, 'plain'))

            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=10) as server:
                if self.use_tls:
                    server.starttls()
                server.login(self.email_address, self.password)
                server.send_message(msg)

            logger.info(f"Correo enviado a {to_email}")
            return True, "Correo enviado exitosamente"

        except socket.timeout:
            error_msg = f"Timeout al enviar correo a {to_email}"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            logger.error(f"Error al enviar correo a {to_email}: {e}")
            return False, str(e)

    def send_email_with_attachment(self, to_email, subject, body, attachment_path):
        """
        Envía un correo con un archivo adjunto.

        Args:
            to_email: Destinatario del correo
            subject: Asunto del correo
            body: Cuerpo del mensaje
            attachment_path: Ruta completa del archivo a adjuntar

        Returns:
            tuple: (success, message)
        """
        try:
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart

            msg = MIMEMultipart()
            msg['From'] = self.email_address
            msg['To'] = to_email
            msg['Subject'] = subject

            # Adjuntar cuerpo del mensaje
            msg.attach(MIMEText(body, 'plain'))

            # Adjuntar archivo
            if attachment_path and os.path.exists(attachment_path):
                filename = os.path.basename(attachment_path)

                with open(attachment_path, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())

                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {filename}'
                )
                msg.attach(part)

                logger.info(f"Archivo adjunto agregado: {filename}")

            # Enviar correo
            with smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=10) as server:
                if self.use_tls:
                    server.starttls()
                server.login(self.email_address, self.password)
                server.send_message(msg)

            logger.info(f"Correo con adjunto enviado a {to_email}")
            return True, "Correo con adjunto enviado exitosamente"

        except socket.timeout:
            error_msg = f"Timeout al enviar correo a {to_email}"
            logger.error(error_msg)
            return False, error_msg
        except Exception as e:
            logger.error(f"Error al enviar correo con adjunto a {to_email}: {e}")
            return False, str(e)

    @staticmethod
    def extract_excel_data(excel_path, row=1, col_g='G', col_h='H'):
        """
        Extrae datos de un archivo Excel en posiciones específicas.
        NOTA: Este método se mantiene por compatibilidad. Para IMEIs usar extract_all_imeis_from_excel.

        Args:
            excel_path: Ruta del archivo Excel
            row: Número de fila (default: 1)
            col_g: Columna para el primer dato (default: 'G')
            col_h: Columna para el segundo dato (default: 'H')

        Returns:
            dict: {
                'success': bool,
                'data_g': valor de columna G,
                'data_h': valor de columna H,
                'error': mensaje de error (si aplica)
            }
        """
        result = {
            'success': False,
            'data_g': None,
            'data_h': None,
            'error': None
        }

        try:
            import openpyxl

            # Verificar que el archivo existe
            if not os.path.exists(excel_path):
                result['error'] = f"Archivo no encontrado: {excel_path}"
                logger.error(result['error'])
                return result

            # Abrir el archivo Excel
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active

            # Extraer datos de las columnas G y H en la fila especificada
            cell_g = f"{col_g}{row}"
            cell_h = f"{col_h}{row}"

            value_g = sheet[cell_g].value
            value_h = sheet[cell_h].value

            result['data_g'] = value_g
            result['data_h'] = value_h
            result['success'] = True

            workbook.close()

            logger.info(f"Datos extraídos del Excel: G{row}={value_g}, H{row}={value_h}")

        except ImportError:
            result['error'] = "La librería 'openpyxl' no está instalada. Ejecute: pip install openpyxl"
            logger.error(result['error'])
        except Exception as e:
            result['error'] = f"Error al extraer datos del Excel: {str(e)}"
            logger.error(result['error'])

        return result

    @staticmethod
    def extract_all_imeis_from_excel(excel_path):
        """
        Extrae todos los IMEIs y fechas de un archivo Excel.
        Lee desde la fila 2 (asumiendo fila 1 son encabezados):
        - Columna A: IMEI
        - Columna B: Registered at (fecha_cliente)

        Args:
            excel_path: Ruta del archivo Excel

        Returns:
            dict: {
                'success': bool,
                'data': lista de diccionarios con 'imei' y 'fecha_cliente',
                'total_rows': número total de filas procesadas,
                'error': mensaje de error (si aplica)
            }
        """
        result = {
            'success': False,
            'data': [],
            'total_rows': 0,
            'error': None
        }

        try:
            import openpyxl
            from datetime import datetime

            # Verificar que el archivo existe
            if not os.path.exists(excel_path):
                result['error'] = f"Archivo no encontrado: {excel_path}"
                logger.error(result['error'])
                return result

            # Abrir el archivo Excel
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active

            # Procesar desde la fila 2 (fila 1 son encabezados)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # row[0] = columna A (IMEI)
                # row[1] = columna B (Registered at)

                imei = row[0] if len(row) > 0 else None
                fecha_raw = row[1] if len(row) > 1 else None

                # Saltar filas vacías
                if not imei:
                    continue

                # Convertir imei a string y limpiar
                imei_str = str(imei).strip()
                if not imei_str:
                    continue

                # Procesar la fecha
                fecha_cliente = None
                if fecha_raw:
                    # Si ya es un objeto datetime
                    if isinstance(fecha_raw, datetime):
                        fecha_cliente = fecha_raw
                    # Si es un string, intentar parsearlo
                    elif isinstance(fecha_raw, str):
                        try:
                            # Intentar varios formatos comunes
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                                try:
                                    fecha_cliente = datetime.strptime(fecha_raw.strip(), fmt)
                                    break
                                except ValueError:
                                    continue
                        except Exception as e:
                            logger.warning(f"No se pudo parsear fecha '{fecha_raw}': {e}")

                # Agregar a la lista
                result['data'].append({
                    'imei': imei_str,
                    'fecha_cliente': fecha_cliente
                })
                result['total_rows'] += 1

            workbook.close()

            result['success'] = True
            logger.info(f"Extraídos {result['total_rows']} IMEIs del Excel: {excel_path}")

        except ImportError:
            result['error'] = "La librería 'openpyxl' no está instalada. Ejecute: pip install openpyxl"
            logger.error(result['error'])
        except Exception as e:
            result['error'] = f"Error al extraer IMEIs del Excel: {str(e)}"
            logger.error(result['error'])
            import traceback
            logger.debug(traceback.format_exc())

        return result

    @staticmethod
    def create_text_file_with_data(data_g, data_h, excel_filename, output_dir=None):
        """
        Crea un archivo de texto con los datos extraídos del Excel.

        Args:
            data_g: Dato de la columna G
            data_h: Dato de la columna H
            excel_filename: Nombre del archivo Excel procesado
            output_dir: Directorio donde guardar el archivo (default: temp)

        Returns:
            tuple: (success, file_path, error_message)
        """
        try:
            # Definir directorio de salida
            if output_dir is None:
                output_dir = tempfile.gettempdir()

            # Crear nombre de archivo con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"datos_extraidos_{timestamp}.txt"
            filepath = os.path.join(output_dir, filename)

            # Crear contenido del archivo
            content = f"""=== DATOS EXTRAÍDOS DEL EXCEL ===
Fecha de extracción: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Columna G (Fila 1): {data_g if data_g is not None else 'Sin datos'}
Columna H (Fila 1): {data_h if data_h is not None else 'Sin datos'}

Archivo procesado: {excel_filename}

---
Generado automáticamente por BotLibertyBD
"""

            # Escribir archivo
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)

            logger.info(f"Archivo de texto creado: {filepath}")
            return True, filepath, None

        except Exception as e:
            error_msg = f"Error al crear archivo de texto: {str(e)}"
            logger.error(error_msg)
            return False, None, error_msg

    @staticmethod
    def analizar_cambios_bd(excel_data, postgres_connector, schema, table):
        """
        Analiza qué cambios se realizarán en la BD antes de sincronizar.
        Compara los IMEIs del Excel con los existentes en la base de datos.

        Args:
            excel_data: Lista de diccionarios con 'imei' y 'fecha_cliente' del Excel
            postgres_connector: Conector de PostgreSQL
            schema: Esquema de la base de datos
            table: Tabla de la base de datos

        Returns:
            dict: {
                'success': bool,
                'nuevos': lista de IMEIs que se insertarán (no existen en BD),
                'actualizados': lista de IMEIs que se actualizarán (existen pero con datos diferentes),
                'sin_cambios': lista de IMEIs que ya existen con los mismos datos,
                'total': total de registros procesados,
                'error': mensaje de error (si aplica)
            }
        """
        resultado = {
            'success': False,
            'nuevos': [],
            'actualizados': [],
            'sin_cambios': [],
            'total': 0,
            'error': None
        }

        try:
            if not postgres_connector or not excel_data:
                resultado['error'] = "Conector PostgreSQL o datos del Excel no disponibles"
                return resultado

            # Extraer solo los IMEIs del Excel para consulta
            imeis_excel = [item['imei'] for item in excel_data]
            resultado['total'] = len(imeis_excel)

            if not imeis_excel:
                resultado['success'] = True
                return resultado

            # Consultar IMEIs existentes en la base de datos
            query = f"""
                SELECT imei, fecha_cliente, activo
                FROM {schema}.{table}
                WHERE imei = ANY(%s)
            """

            existing_imeis = postgres_connector.execute_query(query, (imeis_excel,))

            # Validar que execute_query no retornó None
            if existing_imeis is None:
                resultado['error'] = "Error al consultar BD: No se pudo ejecutar la consulta (conexión perdida)"
                logger.error("execute_query retornó None - posible pérdida de conexión PostgreSQL")
                return resultado

            if not existing_imeis.get('success', False):
                resultado['error'] = f"Error al consultar BD: {', '.join(existing_imeis.get('errors', []))}"
                return resultado

            # Crear diccionario de IMEIs existentes para búsqueda rápida
            existing_dict = {}
            if existing_imeis.get('data'):
                for row in existing_imeis['data']:
                    imei = row[0]
                    fecha_cliente = row[1]
                    activo = row[2]
                    existing_dict[imei] = {'fecha_cliente': fecha_cliente, 'activo': activo}

            # Clasificar cada IMEI del Excel
            for item in excel_data:
                imei = item['imei']
                fecha_cliente = item.get('fecha_cliente')

                if imei not in existing_dict:
                    # IMEI no existe en BD -> será insertado
                    resultado['nuevos'].append({
                        'imei': imei,
                        'fecha_cliente': fecha_cliente
                    })
                else:
                    # IMEI existe -> verificar si cambiará
                    existing_fecha = existing_dict[imei]['fecha_cliente']
                    existing_activo = existing_dict[imei]['activo']

                    # Comparar fechas (considerar None como equivalente)
                    fechas_diferentes = False
                    if fecha_cliente is not None and existing_fecha is not None:
                        # Comparar fechas normalizando a solo fecha (sin hora)
                        if hasattr(fecha_cliente, 'date'):
                            fecha_excel = fecha_cliente.date()
                        else:
                            fecha_excel = fecha_cliente

                        if hasattr(existing_fecha, 'date'):
                            fecha_bd = existing_fecha.date()
                        else:
                            fecha_bd = existing_fecha

                        fechas_diferentes = fecha_excel != fecha_bd
                    elif fecha_cliente != existing_fecha:
                        fechas_diferentes = True

                    # Si la fecha cambió o el registro estaba inactivo, será actualizado
                    if fechas_diferentes or not existing_activo:
                        resultado['actualizados'].append({
                            'imei': imei,
                            'fecha_cliente': fecha_cliente,
                            'fecha_anterior': existing_fecha,
                            'estaba_inactivo': not existing_activo
                        })
                    else:
                        # Sin cambios
                        resultado['sin_cambios'].append({
                            'imei': imei,
                            'fecha_cliente': fecha_cliente
                        })

            resultado['success'] = True
            logger.info(f"Análisis completado: {len(resultado['nuevos'])} nuevos, "
                       f"{len(resultado['actualizados'])} actualizados, "
                       f"{len(resultado['sin_cambios'])} sin cambios")

        except Exception as e:
            resultado['error'] = f"Error al analizar cambios: {str(e)}"
            logger.error(resultado['error'])
            import traceback
            logger.debug(traceback.format_exc())

        return resultado

    @staticmethod
    def generar_reporte_pdf(analisis_datos, archivo_excel, ruta_salida, sync_result=None):
        """
        Genera un reporte en PDF con el análisis de cambios en la BD.

        Args:
            analisis_datos: Resultado de analizar_cambios_bd()
            archivo_excel: Nombre del archivo Excel procesado
            ruta_salida: Ruta donde guardar el PDF
            sync_result: Resultado de sync_imeis() con información de desactivados (opcional)

        Returns:
            tuple: (success, file_path, error_message)
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            import matplotlib
            matplotlib.use('Agg')  # Backend sin GUI
            import matplotlib.pyplot as plt
            import tempfile

            # Crear PDF
            pdf_path = os.path.join(ruta_salida, 'reporte_sincronizacion.pdf')
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            # Estilo personalizado para título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a237e'),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            # Estilo para subtítulos
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#283593'),
                spaceAfter=12,
                spaceBefore=12
            )

            # ===== ENCABEZADO =====
            story.append(Paragraph("BotLibertyBD", title_style))
            story.append(Paragraph("Reporte de Sincronización de IMEIs", subtitle_style))
            story.append(Spacer(1, 0.2 * inch))

            # Información general
            fecha_generacion = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            info_data = [
                ['Fecha de Generación:', fecha_generacion],
                ['Archivo Procesado:', archivo_excel],
            ]

            info_table = Table(info_data, colWidths=[2.5*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(info_table)
            story.append(Spacer(1, 0.3 * inch))

            # ===== EXTRAER DATOS =====
            total_registros = analisis_datos.get('total', 0)
            nuevos = len(analisis_datos.get('nuevos', []))
            actualizados = len(analisis_datos.get('actualizados', []))
            sin_cambios = len(analisis_datos.get('sin_cambios', []))

            # Extraer desactivados del sync_result si está disponible
            desactivados = 0
            if sync_result and isinstance(sync_result, dict):
                desactivados = sync_result.get('desactivados', 0)

            # ===== GENERAR GRÁFICO DE BARRAS =====
            try:
                # Crear figura para el gráfico
                fig, ax = plt.subplots(figsize=(8, 5))

                # Datos para el gráfico
                categorias = ['Nuevos', 'Actualizados', 'Desactivados', 'Sin Cambios']
                valores = [nuevos, actualizados, desactivados, sin_cambios]
                colores_barras = ['#4caf50', '#ff9800', '#f44336', '#9e9e9e']

                # Crear gráfico de barras
                barras = ax.bar(categorias, valores, color=colores_barras, edgecolor='black', linewidth=1.2)

                # Agregar valores encima de las barras
                for barra in barras:
                    altura = barra.get_height()
                    ax.text(barra.get_x() + barra.get_width()/2., altura,
                           f'{int(altura)}',
                           ha='center', va='bottom', fontsize=11, fontweight='bold')

                # Configuración del gráfico
                ax.set_ylabel('Cantidad de IMEIs', fontsize=12, fontweight='bold')
                ax.set_title('Cambios en Base de Datos', fontsize=14, fontweight='bold', pad=20)
                ax.set_ylim(0, max(valores) * 1.15 if max(valores) > 0 else 10)
                ax.grid(axis='y', alpha=0.3, linestyle='--')

                # Mejorar apariencia
                plt.tight_layout()

                # Guardar gráfico en archivo temporal
                grafico_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.png', dir=ruta_salida)
                plt.savefig(grafico_temp.name, dpi=150, bbox_inches='tight')
                plt.close(fig)

                # Agregar gráfico al PDF
                story.append(Paragraph("📊 Resumen Visual de Cambios", subtitle_style))
                img = Image(grafico_temp.name, width=6*inch, height=3.75*inch)
                story.append(img)
                story.append(Spacer(1, 0.3 * inch))

                # Eliminar archivo temporal del gráfico
                import os as os_module
                try:
                    os_module.unlink(grafico_temp.name)
                except:
                    pass

            except Exception as e:
                logger.warning(f"No se pudo generar gráfico: {str(e)}")
                # Continuar sin el gráfico si hay error

            # ===== ESTADÍSTICAS =====
            story.append(Paragraph("📊 Estadísticas Detalladas", subtitle_style))

            stats_data = [
                ['Métrica', 'Cantidad', 'Porcentaje'],
                ['Total de Registros en Excel', str(total_registros), '100%'],
                ['📥 Nuevos (INSERT)', str(nuevos), f'{(nuevos/total_registros*100):.1f}%' if total_registros > 0 else '0%'],
                ['🔄 Actualizados (UPDATE)', str(actualizados), f'{(actualizados/total_registros*100):.1f}%' if total_registros > 0 else '0%'],
                ['🚫 Desactivados (ya no en Excel)', str(desactivados), 'N/A'],
                ['✓ Sin Cambios', str(sin_cambios), f'{(sin_cambios/total_registros*100):.1f}%' if total_registros > 0 else '0%'],
            ]

            stats_table = Table(stats_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e8eaf6')),
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#c5e1a5')),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fff9c4')),
                ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#ffcdd2')),
                ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#e0e0e0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3 * inch))

            # ===== ANÁLISIS =====
            story.append(Paragraph("📈 Análisis de Cambios", subtitle_style))

            analisis_text = ""
            if total_registros > 0:
                porcentaje_nuevos = (nuevos/total_registros*100)
                porcentaje_actualizados = (actualizados/total_registros*100)
                porcentaje_sin_cambios = (sin_cambios/total_registros*100)

                # Análisis de nuevos
                if porcentaje_nuevos >= 50:
                    analisis_text += f"• <b>Crecimiento significativo:</b> El {porcentaje_nuevos:.1f}% de los registros son nuevos ({nuevos} IMEIs agregados a la BD).<br/>"
                elif nuevos > 0:
                    analisis_text += f"• Se agregaron {nuevos} nuevos IMEIs ({porcentaje_nuevos:.1f}% del total).<br/>"

                # Análisis de actualizados
                if porcentaje_actualizados >= 30:
                    analisis_text += f"• <b>Actualización masiva:</b> {actualizados} IMEIs fueron actualizados ({porcentaje_actualizados:.1f}%).<br/>"
                elif actualizados > 0:
                    analisis_text += f"• Se actualizaron {actualizados} IMEIs existentes ({porcentaje_actualizados:.1f}%).<br/>"

                # Análisis de desactivados (NUEVO)
                if desactivados > 0:
                    analisis_text += f"• <b>IMEIs desactivados:</b> {desactivados} IMEIs ya no aparecen en el archivo Excel (marcados como inactivos en BD).<br/>"

                # Análisis de sin cambios
                if porcentaje_sin_cambios >= 50:
                    analisis_text += f"• Base estable: {porcentaje_sin_cambios:.1f}% de los registros ya existían sin cambios.<br/>"

                # Resumen general si no hay análisis específico
                if not analisis_text:
                    analisis_text = f"• Se procesaron {total_registros} registros del Excel con {nuevos} nuevos, {actualizados} actualizaciones"
                    if desactivados > 0:
                        analisis_text += f" y {desactivados} desactivados"
                    analisis_text += ".<br/>"
            else:
                analisis_text = "• No se procesaron registros."

            story.append(Paragraph(analisis_text, styles['BodyText']))
            story.append(Spacer(1, 0.3 * inch))

            # ===== DETALLES DE REGISTROS =====
            # Mostrar primeros 10 nuevos (si hay)
            if nuevos > 0:
                story.append(Paragraph("📥 Registros Nuevos (Primeros 10)", subtitle_style))
                detalle_nuevos = [['#', 'IMEI', 'Fecha Cliente']]

                for idx, item in enumerate(analisis_datos['nuevos'][:10], 1):
                    imei = item['imei']
                    fecha = item.get('fecha_cliente', 'N/A')
                    if fecha and hasattr(fecha, 'strftime'):
                        fecha = fecha.strftime('%Y-%m-%d')
                    detalle_nuevos.append([str(idx), str(imei), str(fecha) if fecha else 'N/A'])

                if nuevos > 10:
                    detalle_nuevos.append(['...', f'(+{nuevos - 10} más)', '...'])

                nuevos_table = Table(detalle_nuevos, colWidths=[0.5*inch, 3*inch, 2.5*inch])
                nuevos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4caf50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f8e9')]),
                ]))
                story.append(nuevos_table)
                story.append(Spacer(1, 0.2 * inch))

            # Mostrar primeros 10 actualizados (si hay)
            if actualizados > 0:
                story.append(Paragraph("🔄 Registros Actualizados (Primeros 10)", subtitle_style))
                detalle_actualizados = [['#', 'IMEI', 'Fecha Nueva', 'Fecha Anterior']]

                for idx, item in enumerate(analisis_datos['actualizados'][:10], 1):
                    imei = item['imei']
                    fecha_nueva = item.get('fecha_cliente', 'N/A')
                    fecha_anterior = item.get('fecha_anterior', 'N/A')

                    if fecha_nueva and hasattr(fecha_nueva, 'strftime'):
                        fecha_nueva = fecha_nueva.strftime('%Y-%m-%d')
                    if fecha_anterior and hasattr(fecha_anterior, 'strftime'):
                        fecha_anterior = fecha_anterior.strftime('%Y-%m-%d')

                    detalle_actualizados.append([
                        str(idx),
                        str(imei),
                        str(fecha_nueva) if fecha_nueva else 'N/A',
                        str(fecha_anterior) if fecha_anterior else 'N/A'
                    ])

                if actualizados > 10:
                    detalle_actualizados.append(['...', f'(+{actualizados - 10} más)', '...', '...'])

                actualizados_table = Table(detalle_actualizados, colWidths=[0.5*inch, 2.5*inch, 1.5*inch, 1.5*inch])
                actualizados_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ff9800')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff3e0')]),
                ]))
                story.append(actualizados_table)
                story.append(Spacer(1, 0.2 * inch))

            # ===== PIE DE PÁGINA =====
            story.append(Spacer(1, 0.5 * inch))
            timestamp_final = datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC-6')
            footer_text = f"<br/><br/>---<br/>[Timestamp: {timestamp_final}]<br/>Generado automáticamente por BotLibertyBD"
            story.append(Paragraph(footer_text, styles['Normal']))

            # Construir PDF
            doc.build(story)

            logger.info(f"Reporte PDF generado: {pdf_path}")
            return True, pdf_path, None

        except ImportError as e:
            error_msg = f"Error: La librería 'reportlab' no está instalada. Ejecute: pip install reportlab"
            logger.error(error_msg)
            return False, None, error_msg
        except Exception as e:
            error_msg = f"Error al generar PDF: {str(e)}"
            logger.error(error_msg)
            import traceback
            logger.debug(traceback.format_exc())
            return False, None, error_msg

    def monitor_and_notify(self, title_filter, notify_emails, folder_path="INBOX",
                          status_callback=None, max_emails_to_check=50, max_matches=5,
                          postgres_connector=None, schema="automatizacion", table="datos_excel_doforms"):
        """
        Monitorea correos no leídos con un título específico, los marca como leídos
        y envía notificaciones a los usuarios especificados.
        OPTIMIZADO: Procesa correos uno a uno con límites para evitar consumo excesivo.

        Args:
            title_filter: Filtro de título para buscar correos
            notify_emails: Lista de emails a notificar
            folder_path: Carpeta de correo a monitorear (default: INBOX)
            status_callback: Callback para reportar estado
            max_emails_to_check: Máximo de correos a revisar por ciclo (default: 50)
            max_matches: Máximo de coincidencias a procesar por ciclo (default: 5)
            postgres_connector: Conector de PostgreSQL para sincronizar IMEIs (opcional)
            schema: Esquema de la base de datos (default: automatizacion)
            table: Tabla de la base de datos (default: datos_excel_doforms)
        """
        results = {
            "success": False,
            "total_items": 0,
            "matching_items": 0,
            "notified_users": 0,
            "imeis_sincronizados": 0,
            "errors": [],
            "message": ""
        }

        folder_path = (folder_path or "INBOX").strip() or "INBOX"
        prepared_filters = self._prepare_title_filters(title_filter)

        imap = None
        try:
            # Agregar timeout de 30 segundos para evitar bloqueos indefinidos
            imap = imaplib.IMAP4_SSL(self.imap_server, self.imap_port, timeout=30)
            imap.login(self.email_address, self.password)
            imap.select(folder_path)

            # Buscar solo correos NO LEÍDOS de hoy
            date_str = datetime.now().strftime('%d-%b-%Y')
            typ, data = imap.search(None, 'UNSEEN', 'ON', date_str)

            if typ != 'OK':
                results["message"] = "Error en búsqueda IMAP"
                return results

            email_ids = data[0].split()
            total_unread = len(email_ids)

            if not email_ids:
                results["success"] = True
                results["message"] = "No hay correos nuevos"
                return results

            # OPTIMIZACIÓN: Limitar cantidad de correos a revisar
            emails_to_process = email_ids[:max_emails_to_check]

            if status_callback:
                if total_unread > max_emails_to_check:
                    status_callback(f"Encontrados {total_unread} correos no leídos. Procesando los primeros {max_emails_to_check}...", "INFO")
                else:
                    status_callback(f"Encontrados {total_unread} correos no leídos. Procesando...", "INFO")

            # Procesar cada correo UNO A UNO
            emails_checked = 0
            for num in emails_to_process:
                emails_checked += 1

                # OPTIMIZACIÓN: Detener si ya encontramos suficientes coincidencias
                if results["matching_items"] >= max_matches:
                    if status_callback:
                        status_callback(f"Límite de {max_matches} coincidencias alcanzado. Deteniendo búsqueda.", "INFO")
                    break

                # Leer encabezado para obtener asunto (uno a uno)
                typ, header_data = imap.fetch(num, '(BODY.PEEK[HEADER.FIELDS (SUBJECT FROM)])')
                if typ != 'OK':
                    continue

                header_msg = email.message_from_bytes(header_data[0][1])
                subject = header_msg.get('Subject', '')
                from_email = header_msg.get('From', '')

                # Verificar si el asunto coincide con el filtro
                if not self._subject_matches(subject, prepared_filters):
                    # No coincide, continuar con el siguiente
                    if status_callback and emails_checked % 10 == 0:
                        status_callback(f"Revisados {emails_checked} correos...", "INFO")
                    continue

                # COINCIDENCIA ENCONTRADA
                results["matching_items"] += 1

                if status_callback:
                    status_callback(f"✓ Coincidencia #{results['matching_items']}: '{subject}' de {from_email}", "SUCCESS")

                # Descargar el mensaje completo para obtener adjuntos
                typ, full_msg_data = imap.fetch(num, '(RFC822)')
                if typ != 'OK':
                    if status_callback:
                        status_callback("Error al descargar mensaje completo", "WARNING")
                    continue

                full_msg = email.message_from_bytes(full_msg_data[0][1])

                # Variables para almacenar archivos temporales
                excel_files = []
                text_file_path = None
                temp_dir = tempfile.mkdtemp(prefix="bot_liberty_")

                try:
                    # Buscar y descargar adjuntos Excel
                    for part in full_msg.walk():
                        if part.get_content_disposition() == 'attachment':
                            filename = part.get_filename()
                            if filename and filename.lower().endswith(('.xls', '.xlsx')):
                                filepath = os.path.join(temp_dir, filename)
                                with open(filepath, 'wb') as f:
                                    f.write(part.get_payload(decode=True))
                                excel_files.append(filepath)
                                if status_callback:
                                    status_callback(f"📎 Excel descargado: {filename}", "INFO")

                    # Procesar el primer archivo Excel encontrado
                    if excel_files:
                        excel_path = excel_files[0]
                        excel_filename = os.path.basename(excel_path)

                        if status_callback:
                            status_callback(f"📊 Extrayendo IMEIs del Excel...", "INFO")

                        # Extraer todos los IMEIs del Excel (columnas A y B)
                        extraction_result = self.extract_all_imeis_from_excel(excel_path)

                        if extraction_result['success']:
                            total_imeis = extraction_result['total_rows']

                            if status_callback:
                                status_callback(f"✓ {total_imeis} IMEI(s) extraídos del Excel", "SUCCESS")

                            # Variables para almacenar el análisis
                            analisis_datos = None
                            sync_result = None

                            # Analizar cambios en la BD ANTES de sincronizar (si hay conector PostgreSQL)
                            if postgres_connector and total_imeis > 0:
                                if status_callback:
                                    status_callback(f"🔍 Analizando cambios en la base de datos...", "INFO")

                                analisis_datos = self.analizar_cambios_bd(
                                    excel_data=extraction_result['data'],
                                    postgres_connector=postgres_connector,
                                    schema=schema,
                                    table=table
                                )

                                # Validar que analizar_cambios_bd retornó un resultado válido
                                if analisis_datos and analisis_datos.get('success', False):
                                    nuevos_count = len(analisis_datos.get('nuevos', []))
                                    actualizados_count = len(analisis_datos.get('actualizados', []))
                                    sin_cambios_count = len(analisis_datos.get('sin_cambios', []))

                                    if status_callback:
                                        status_callback(
                                            f"✓ Análisis completado: {nuevos_count} nuevos, "
                                            f"{actualizados_count} actualizados, "
                                            f"{sin_cambios_count} sin cambios",
                                            "SUCCESS"
                                        )
                                elif analisis_datos and not analisis_datos.get('success', False):
                                    if status_callback:
                                        status_callback(f"⚠ Error en análisis: {analisis_datos.get('error', 'Desconocido')}", "WARNING")
                                elif analisis_datos is None:
                                    if status_callback:
                                        status_callback(f"⚠ Error en análisis: Error al analizar cambios (None)", "WARNING")

                                # Sincronizar con la base de datos
                                if status_callback:
                                    status_callback(f"🔄 Sincronizando {total_imeis} IMEIs con la base de datos...", "INFO")

                                sync_result = postgres_connector.sync_imeis(
                                    schema=schema,
                                    table=table,
                                    excel_data=extraction_result['data']
                                )

                                if sync_result['success']:
                                    results['imeis_sincronizados'] = total_imeis
                                    if status_callback:
                                        status_callback(
                                            f"✓ Sincronización completada: {sync_result['nuevos']} nuevos, "
                                            f"{sync_result['actualizados']} actualizados, "
                                            f"{sync_result['desactivados']} desactivados",
                                            "SUCCESS"
                                        )
                                else:
                                    error_msg = f"Error en sincronización: {', '.join(sync_result['errors'])}"
                                    results['errors'].append(error_msg)
                                    if status_callback:
                                        status_callback(f"⚠ {error_msg}", "WARNING")

                            # Generar reporte PDF con análisis (si se pudo analizar)
                            pdf_file_path = None
                            if total_imeis > 0 and analisis_datos and analisis_datos.get('success', False):
                                if status_callback:
                                    status_callback(f"📄 Generando reporte PDF...", "INFO")

                                pdf_success, pdf_path, pdf_error = self.generar_reporte_pdf(
                                    analisis_datos=analisis_datos,
                                    archivo_excel=excel_filename,
                                    ruta_salida=temp_dir,
                                    sync_result=sync_result
                                )

                                if pdf_success:
                                    pdf_file_path = pdf_path
                                    if status_callback:
                                        status_callback(f"✓ Reporte PDF generado exitosamente", "SUCCESS")
                                else:
                                    if status_callback:
                                        status_callback(f"⚠ Error al generar PDF: {pdf_error}", "WARNING")
                                    logger.warning(f"No se pudo generar PDF: {pdf_error}")

                            # Si no se generó PDF, crear archivo de texto de respaldo (fallback)
                            if not pdf_file_path and total_imeis > 0:
                                summary_content = f"""=== RESUMEN DE PROCESAMIENTO DE IMEIs ===
Fecha de procesamiento: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Total de IMEIs procesados: {total_imeis}

Archivo procesado: {excel_filename}

---
Generado automáticamente por BotLibertyBD
"""
                                summary_file = os.path.join(temp_dir, 'resumen_imeis.txt')
                                with open(summary_file, 'w', encoding='utf-8') as f:
                                    f.write(summary_content)
                                text_file_path = summary_file

                                if status_callback:
                                    status_callback(f"✓ Archivo de resumen creado (fallback)", "INFO")
                        else:
                            if status_callback:
                                status_callback(f"⚠ Error al extraer datos: {extraction_result['error']}", "WARNING")

                    # Marcar como leído después de procesar
                    imap.store(num, '+FLAGS', '\\Seen')

                    # Enviar notificación a cada usuario
                    for notify_email in notify_emails:
                        # Preparar timestamp para el título
                        timestamp_actual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        notification_subject = f"Notificación de Procesamiento - BotLibertyBD {timestamp_actual}"

                        notification_body = "Se ha detectado y procesado exitosamente un correo con archivos adjuntos.\n\n"

                        # Agregar resumen de procesamiento si hay datos de análisis
                        if analisis_datos and analisis_datos.get('success', False):
                            nuevos_count = len(analisis_datos.get('nuevos', []))
                            actualizados_count = len(analisis_datos.get('actualizados', []))
                            sin_cambios_count = len(analisis_datos.get('sin_cambios', []))
                            total_count = analisis_datos.get('total', 0)

                            notification_body += "📊 RESUMEN DE PROCESAMIENTO:\n"
                            notification_body += f"• Total de IMEIs sincronizados: {total_count}\n"
                            notification_body += f"• Registros nuevos agregados: {nuevos_count}\n"
                            notification_body += f"• Registros actualizados: {actualizados_count}\n"
                            notification_body += f"• Registros sin cambios: {sin_cambios_count}\n"

                            if excel_files:
                                notification_body += f"• Archivo procesado: {os.path.basename(excel_files[0])}\n"
                        elif excel_files:
                            # Si no hay análisis pero sí hay archivos Excel
                            notification_body += "📊 RESUMEN DE PROCESAMIENTO:\n"
                            notification_body += f"• Archivo procesado: {os.path.basename(excel_files[0])}\n"

                        # Indicar si hay PDF adjunto
                        if pdf_file_path and os.path.exists(pdf_file_path):
                            notification_body += "\n📎 Se adjunta un reporte detallado en formato PDF con el análisis completo.\n"
                        elif text_file_path and os.path.exists(text_file_path):
                            notification_body += "\n📎 Se adjunta un archivo de resumen con los datos procesados.\n"

                        # Pie de mensaje
                        notification_body += "\n---\nEste es un mensaje automático de BotLibertyBD."

                        # Enviar correo con adjunto (preferir PDF, luego TXT, sino sin adjunto)
                        attachment_to_send = None
                        if pdf_file_path and os.path.exists(pdf_file_path):
                            attachment_to_send = pdf_file_path
                        elif text_file_path and os.path.exists(text_file_path):
                            attachment_to_send = text_file_path

                        if attachment_to_send:
                            success, message = self.send_email_with_attachment(
                                notify_email,
                                notification_subject,
                                notification_body,
                                attachment_to_send
                            )
                        else:
                            success, message = self.send_simple_email(
                                notify_email,
                                notification_subject,
                                notification_body
                            )

                        if success:
                            results["notified_users"] += 1
                            if status_callback:
                                status_callback(f"✉ Notificación enviada a {notify_email}", "SUCCESS")
                        else:
                            error_msg = f"Error al notificar a {notify_email}: {message}"
                            results["errors"].append(error_msg)
                            if status_callback:
                                status_callback(error_msg, "ERROR")

                finally:
                    # Limpiar archivos temporales
                    try:
                        import shutil
                        if os.path.exists(temp_dir):
                            shutil.rmtree(temp_dir)
                            if status_callback:
                                status_callback(f"🗑 Archivos temporales eliminados", "INFO")
                    except Exception as e:
                        logger.warning(f"No se pudieron eliminar archivos temporales: {e}")

            results["success"] = True
            results["total_items"] = emails_checked
            results["message"] = f"{results['matching_items']} correo(s) detectado(s), {results['notified_users']} notificación(es) enviada(s)"

        except socket.timeout:
            error_msg = "Timeout de conexión IMAP en monitoreo"
            results["errors"].append(error_msg)
            results["message"] = error_msg
            if status_callback:
                status_callback(error_msg, "ERROR")
            logger.error(error_msg)
        except Exception as e:
            error_msg = f"Error en monitoreo: {str(e)}"
            results["errors"].append(error_msg)
            results["message"] = error_msg
            if status_callback:
                status_callback(error_msg, "ERROR")
            logger.error(error_msg)
        finally:
            if imap:
                try:
                    imap.logout()
                except:
                    pass

        return results
